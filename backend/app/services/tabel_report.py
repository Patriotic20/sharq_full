"""TABEL monthly timesheet — data builder + Excel (.xlsx) renderer."""
from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date as DateType, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.timezone import APP_TZ
from app.enums.attendance import AttendanceStatus, LeaveType
from app.models.attendances import Attendance
from app.models.employees import Employee
from app.models.holiday import Holiday
from app.repositories.holiday import HolidayRepository


ORG_NAME = "SHARQ-UNIVERSITETI"
RECTOR_NAME = "Sh.A. Atamuradov"
DEFAULT_KAFEDRA = "Iqtisodiyot va axborot texnologiyalari"

MONTHS_UZ = [
    "YANVAR", "FEVRAL", "MART", "APREL", "MAY", "IYUN",
    "IYUL", "AVGUST", "SENTABR", "OKTABR", "NOYABR", "DEKABR",
]

LEAVE_TO_CODE: dict[LeaveType, str] = {
    LeaveType.SICK: "F",
    LeaveType.VACATION_ANNUAL: "RP",
    LeaveType.VACATION_EDUCATION: "G",
    LeaveType.LEAVE_EDUCATION: "N",
    LeaveType.ADMIN_ABSENCE: "V",
    LeaveType.STATE_DUTY: "OU",
    LeaveType.MATERNITY: "R",
}

WORKED_CODES = {"B", "O"}
EXCUSED_CODES = {"F", "V", "G", "N", "RP", "R", "OU"}

FILL_HOLIDAY = PatternFill("solid", fgColor="FFE0E0E0")
FILL_WEEKEND = PatternFill("solid", fgColor="FFF2F2F2")
FILL_WORK = PatternFill("solid", fgColor="FFE8F5E9")
FILL_ABSENT_EXCUSED = PatternFill("solid", fgColor="FFFFF8E1")
FILL_TRUANCY = PatternFill("solid", fgColor="FFFFEBEE")

THIN = Side(border_style="thin", color="FFBDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------


@dataclass
class TabelDayCell:
    date: DateType
    code: str
    is_holiday: bool
    is_weekend: bool
    attendance_id: int | None
    status: str | None
    leave_type: str | None


@dataclass
class TabelEmployeeRow:
    employee_id: int
    full_name: str
    position: str | None
    employment_rate: float
    days: list[TabelDayCell]
    worked_days: int


@dataclass
class TabelData:
    year: int
    month: int
    month_name: str
    days_in_month: int
    working_days: int
    org_name: str
    rector_name: str
    kafedra_name: str
    holiday_dates: list[DateType]
    employees: list[TabelEmployeeRow] = field(default_factory=list)


def _local_date(dt: datetime | None) -> DateType | None:
    if dt is None:
        return None
    return dt.astimezone(APP_TZ).date()


def _resolve_holiday_dates(
    year: int,
    month: int,
    holidays: list[Holiday],
    recurring: list[Holiday],
) -> set[DateType]:
    days: set[DateType] = set()
    for h in holidays:
        days.add(h.date)
    for h in recurring:
        try:
            days.add(DateType(year, h.date.month, h.date.day))
        except ValueError:
            continue
    last_day = calendar.monthrange(year, month)[1]
    return {d for d in days if d.year == year and d.month == month and 1 <= d.day <= last_day}


def _code_for_day(
    day: DateType,
    is_holiday: bool,
    is_weekend: bool,
    attendance: Attendance | None,
) -> str:
    if attendance is None:
        if is_holiday or is_weekend:
            return "A"
        return ""
    if attendance.leave_type is not None:
        return LEAVE_TO_CODE.get(attendance.leave_type, "")
    if attendance.status in (AttendanceStatus.PRESENT, AttendanceStatus.LATE, AttendanceStatus.LEFT_EARLY):
        return "O" if is_holiday else "B"
    if attendance.status == AttendanceStatus.ABSENT:
        if is_holiday or is_weekend:
            return "A"
        return "P"
    return ""


async def _load_employees(
    session: AsyncSession,
    department_id: int | None,
) -> list[Employee]:
    q = select(Employee).options(selectinload(Employee.department))
    if department_id is not None:
        q = q.where(Employee.department_id == department_id)
    q = q.order_by(Employee.last_name, Employee.first_name)
    res = await session.execute(q)
    return list(res.scalars().all())


async def _load_attendances(
    session: AsyncSession,
    employee_ids: list[int],
    date_from: DateType,
    date_to: DateType,
) -> dict[tuple[int, DateType], Attendance]:
    if not employee_ids:
        return {}
    from sqlalchemy import func as sa_func
    from app.core.timezone import APP_TZ_NAME

    res = await session.execute(
        select(Attendance).where(
            Attendance.employee_id.in_(employee_ids),
            sa_func.date(sa_func.timezone(APP_TZ_NAME, Attendance.enter_time)) >= date_from,
            sa_func.date(sa_func.timezone(APP_TZ_NAME, Attendance.enter_time)) <= date_to,
        )
    )
    by_key: dict[tuple[int, DateType], Attendance] = {}
    for a in res.scalars().all():
        d = _local_date(a.enter_time) or _local_date(a.exit_time)
        if d is None:
            continue
        by_key.setdefault((a.employee_id, d), a)

    # Second pass for ABSENT records without enter_time (use created_at).
    res2 = await session.execute(
        select(Attendance).where(
            Attendance.employee_id.in_(employee_ids),
            Attendance.enter_time.is_(None),
        )
    )
    for a in res2.scalars().all():
        d = _local_date(a.created_at) if a.created_at else None
        if d is None or not (date_from <= d <= date_to):
            continue
        by_key.setdefault((a.employee_id, d), a)

    return by_key


async def build_tabel_data(
    session: AsyncSession,
    year: int,
    month: int,
    department_id: int | None,
) -> TabelData:
    if not (1 <= month <= 12):
        raise ValueError("month must be 1..12")

    last_day = calendar.monthrange(year, month)[1]
    date_from = DateType(year, month, 1)
    date_to = DateType(year, month, last_day)

    employees = await _load_employees(session, department_id)
    attendances = await _load_attendances(
        session, [e.id for e in employees], date_from, date_to,
    )

    holiday_repo = HolidayRepository(session)
    in_range = await holiday_repo.list_in_range(date_from, date_to)
    recurring = await holiday_repo.list_recurring()
    holiday_days = _resolve_holiday_dates(year, month, in_range, recurring)

    kafedra_name = DEFAULT_KAFEDRA
    if department_id is not None and employees:
        dep = employees[0].department
        if dep and dep.name:
            kafedra_name = dep.name

    working_days = sum(
        1 for d in range(1, last_day + 1)
        if DateType(year, month, d).weekday() < 5 and DateType(year, month, d) not in holiday_days
    )

    rows: list[TabelEmployeeRow] = []
    for emp in employees:
        days: list[TabelDayCell] = []
        worked = 0
        for day_num in range(1, last_day + 1):
            d = DateType(year, month, day_num)
            is_h = d in holiday_days
            is_w = d.weekday() >= 5
            att = attendances.get((emp.id, d))
            code = _code_for_day(d, is_h, is_w, att)
            if code in WORKED_CODES:
                worked += 1
            days.append(TabelDayCell(
                date=d,
                code=code,
                is_holiday=is_h,
                is_weekend=is_w,
                attendance_id=att.id if att is not None else None,
                status=att.status.value if att is not None else None,
                leave_type=att.leave_type.value if (att is not None and att.leave_type is not None) else None,
            ))
        rows.append(TabelEmployeeRow(
            employee_id=emp.id,
            full_name=f"{emp.last_name} {emp.first_name} {emp.middle_name or ''}".strip(),
            position=emp.position.name if emp.position else None,
            employment_rate=float(emp.employment_rate),
            days=days,
            worked_days=worked,
        ))

    return TabelData(
        year=year,
        month=month,
        month_name=MONTHS_UZ[month - 1],
        days_in_month=last_day,
        working_days=working_days,
        org_name=ORG_NAME,
        rector_name=RECTOR_NAME,
        kafedra_name=kafedra_name,
        holiday_dates=sorted(holiday_days),
        employees=rows,
    )


# ---------------------------------------------------------------------------
# Excel renderer
# ---------------------------------------------------------------------------


def _fill_for_code(code: str, is_holiday: bool, is_weekend: bool) -> PatternFill | None:
    if is_holiday:
        return FILL_HOLIDAY
    if code in WORKED_CODES:
        return FILL_WORK
    if code == "P":
        return FILL_TRUANCY
    if code in EXCUSED_CODES:
        return FILL_ABSENT_EXCUSED
    if is_weekend and not code:
        return FILL_WEEKEND
    return None


def _render_xlsx(data: TabelData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "TABEL"

    n_day_cols = data.days_in_month
    total_cols = 4 + n_day_cols + 1

    ws.cell(row=1, column=1, value=f"TASDIQLAYMAN — {data.org_name} rektori {data.rector_name}").font = Font(bold=True)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=total_cols)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=2, column=1, value="TABEL").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=total_cols)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    ws.cell(
        row=3, column=1,
        value=f"Foydalanilgan ish vaqti hisobi va ish haqi hisoblash — {data.kafedra_name} kafedrasi",
    ).font = Font(italic=True)
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=total_cols)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center")

    ws.cell(
        row=4, column=1,
        value=f"{data.year} yil {data.month_name} oyi (ish kunlari soni: {data.working_days})",
    )
    ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=total_cols)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center")

    header_row = 6
    headers = ["№", "F.I.O", "Lavozim", "St."] + [str(d) for d in range(1, n_day_cols + 1)] + ["Jami"]
    holiday_set = set(data.holiday_dates)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if 5 <= col_idx <= 4 + n_day_cols:
            day_num = col_idx - 4
            d = DateType(data.year, data.month, day_num)
            if d in holiday_set:
                cell.fill = FILL_HOLIDAY
            elif d.weekday() >= 5:
                cell.fill = FILL_WEEKEND

    for row_offset, row in enumerate(data.employees, start=1):
        r = header_row + row_offset
        ws.cell(row=r, column=1, value=row_offset).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=row.full_name)
        ws.cell(row=r, column=3, value=row.position or "").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=row.employment_rate).alignment = Alignment(horizontal="center")

        for day_idx, day in enumerate(row.days, start=1):
            cell = ws.cell(row=r, column=4 + day_idx, value=day.code)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            fill = _fill_for_code(day.code, day.is_holiday, day.is_weekend)
            if fill is not None:
                cell.fill = fill
        ws.cell(row=r, column=4 + n_day_cols + 1, value=row.worked_days).font = Font(bold=True)
        ws.cell(row=r, column=4 + n_day_cols + 1).alignment = Alignment(horizontal="center")
        for c in (1, 2, 3, 4, 4 + n_day_cols + 1):
            ws.cell(row=r, column=c).border = BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 6
    for col_idx in range(5, 4 + n_day_cols + 1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = 4
    ws.column_dimensions[ws.cell(row=header_row, column=total_cols).column_letter].width = 8

    legend = wb.create_sheet("Shartli belgilar")
    legend.append(["Kod", "Ko'rsatkichlar nomi"])
    legend["A1"].font = Font(bold=True)
    legend["B1"].font = Font(bold=True)
    for kod, name in (
        ("B", "Haqiqatda ishlangan kunlar"),
        ("O", "Bayramda ishlangan kunlar"),
        ("A", "Dam olish va bayram kunlari"),
        ("F", "Mehnatga layoqatsizlik"),
        ("V", "Ma'muriyat ruxsati bilan ishda qatnashmagan kunlar"),
        ("G", "O'quv ta'tili"),
        ("N", "O'qish bo'yicha dam olishlar"),
        ("RP", "Yillik mehnat ta'tili"),
        ("R", "Tug'ish bilan bog'liq ta'tillar"),
        ("OU", "Davlat oldidagi majburiyatlar bajarish"),
        ("P", "Progullar"),
    ):
        legend.append([kod, name])
    legend.column_dimensions["A"].width = 8
    legend.column_dimensions["B"].width = 55

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_tabel_xlsx(
    session: AsyncSession,
    year: int,
    month: int,
    department_id: int | None,
) -> bytes:
    data = await build_tabel_data(session, year, month, department_id)
    return _render_xlsx(data)
